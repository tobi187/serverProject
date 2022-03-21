import os
from openpyxl import Workbook, load_workbook
import pandas as pd

DATA_SHEET_NAME = "Daten"


class ExcelWorker:
    def __init__(self, save_file_name, base_path):
        self.start_row = 2
        self.col_names = []
        self.save_path = os.path.join(base_path, "uploads", save_file_name)
        self.double_headers = {"Keyword- oder Produkt-Targeting": "Keyword", "Gesamtumsatz für Werbung (ACoS)": "ACOS ", "Verkäufe ": "14 Tage, Umsatz gesamt", "14 Tage, Einheiten gesamt": "Einheiten insgesamt", "Anzeigegruppe ": "Anzeigegruppenname", "SKU ": "Beworbene SKU", "ASIN ": "Beworbene ASIN"}
        self.generate_excel()

    def write_data(self, df: pd.DataFrame):

        for header in df.keys():
            if header in self.double_headers.keys():
                df.rename({header: self.double_headers[header]}, axis=1, inplace=True)

        col_dic = {}
        for header in df.keys():
            if header in self.col_names:
                col_dic[header] = self.col_names.index(header) + 1
            else:
                col_dic[header] = len(self.col_names) + 1
                self.col_names.append(header)

        wb = load_workbook(self.save_path)
        sheet = wb[DATA_SHEET_NAME]

        for col_name, col_index in col_dic.items():
            for row_index, entry in enumerate(df[col_name]):
                sheet.cell(row=row_index + self.start_row, column=col_index).value = entry

        for index, col in enumerate(self.col_names):
            sheet.cell(row=1, column=index + 1).value = col

        self.start_row += len(df[df.keys()[0]]) + 1
        wb.save(self.save_path)

    def generate_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = DATA_SHEET_NAME
        wb.save(self.save_path)
        wb.close()
